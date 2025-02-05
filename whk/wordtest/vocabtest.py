import openpyxl
import random

class vocabtest:
    def __init__(self, filename = "Book1.xlsx"):

        try:
            try : sheet = openpyxl.load_workbook(filename= filename)
            except: sheet = openpyxl.load_workbook(filename= filename + ".xlsx")
        except: sheet = openpyxl.load_workbook(filename="Book1.xlsx")
        sheet = sheet[sheet.sheetnames[0]]

        print("loading vocab...")

        unit = 0
        self.unitdicts = [[] for _ in range(50)]
        for i, _ in enumerate(sheet["A"]):
            for j, cell in enumerate(sheet[str(i+1)]):
                try : 
                    if cell.value.isascii() and cell.value.isalpha():
                        self.unitdicts[unit].append((cell.value, sheet[i+1][j+1].value))
                    else:
                        ix = cell.value.lower().find("unit") 
                        if ix != -1: unit = int(cell.value[ix+4:])
                except: pass
        self.test = None 
        print("done")

    #def gentest(self, units)

    def gentest(self, units:list, num:int, tofile = False):
        totdict = []
        for i in units:
            totdict += self.unitdicts[i]
        random.shuffle(totdict)
        self.test = totdict[:num]
        

    def puttest(self, tofile = False):
        if tofile: 
            output = openpyxl.Workbook()
            output.save("output.xlsx")
            output.create_sheet("Answer")
            output.create_sheet("Question")
        for i, (ur, ul) in enumerate(self.test):
            if tofile: 
                output["Answer"].cell(i//2+1, (i%2)*2+1, ul)
                output["Answer"].cell(i//2+1, (i%2)*2+2, ur)
            print(ul.ljust(7,' '), " \t", ur.ljust(15, ' '), end = "\t\n"[i%2])
        
        print("\n")

        for i, (_, ul) in enumerate(self.test):
            if tofile: 
                output["Question"].cell(i//2+1, (i%2)*2+1, ul)
                #output["Question"].cell(i//2+1, (i%2)*2+2, None)
            print(ul.ljust(7,' '), " \t", " " * 15, end = "\t\n"[i%2])
        
        if tofile: 
            try: output.remove(output["Sheet"])
            except: print("Sheet 'Sheet' is already deleted. How wierd." )
            output.save("output.xlsx")

    def numunit(self):
        return sum([1 if len(ux)!=0 else 0 for ux in self.unitdicts])
    def numword(self):
        return sum([len(ux) for ux in self.unitdicts])

if __name__ == "__main__":
    testi = vocabtest(input("input file name, default = Book1.xlsx"))
    print("input l, r, num")
    l, r, num = map(int, input().split())
    testi.gentest(list(range(l, r+1)), num=num, tofile=False)
    testi.puttest()
